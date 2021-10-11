import pandas as pd
import numpy as np
import openpyxl
import pymysql
from sqlalchemy import create_engine
import os


def init_calss(file):
    df = pd.read_excel(file)[['姓名']]
    df['缺课'] = '0'
    name=file.split("/")[-1]
    df.to_excel(excel_writer="./massage/"+name,index=False,)

def classes(file):
    ls=os.listdir("massage")
    name=file.split("/")[-1]
    if name in ls:
        return 0

def get_class():
    ls = os.listdir("massage")
    return ls


def tongxue(class_id):
    ls=[]
    file="./massage/"+class_id
    df = pd.read_excel(file)[['姓名']]
    data_array=np.array(df)
    data_list=data_array.tolist()
    for i in data_list:
        ls.append(i[0])
    return ls

def update(class_id,id):
    file = "./massage/" + class_id
    wb=openpyxl.load_workbook(file)
    sheet=wb['Sheet1']
    res=sheet.cell(id,2)
    data=int(res.value)+1
    sheet.cell(id,2).value=data
    wb.save(file+"1")
    wb.close()
    os.remove(file)
    os.rename(file+"1",file)
    return data


if __name__=="__main__":
    file="D:/learn/choose_name-master/choose_name/大数据名单.xlsx"
    class_id="大数据2001.xlsx"
    id=24
    #init_calss(file)
    #classes(file)
    print(update(class_id,id))
